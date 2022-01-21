"""
- openpyxl = Se usa para ler planilhas, escrever, criar novas planilhas

.xlsx  =  arquivo excel


"""
import openpyxl
from random import uniform

pedidos = openpyxl.load_workbook('pedidos.xlsx')
nome_planilhas = pedidos.sheetnames
planilha1 = pedidos['Página1']

# APARECE EM FORMA DE TABELA
# É MAIS UTILIZADO PARA LER A PLANILHA, EXTRAIR OS DADOS, JOGAR NUMA BASE DE DADOS, CRIAR LOG, ARQUIVO JSON
for linha in planilha1:
    if linha[0]. value is not None:
        print(linha[0].value, end=" ")
    if linha[1]. value is not None:
        print(linha[1].value, end=" ")
    if linha[2].value is not None:
        print(linha[2].value)


# CRIA NOVOS DADOS, PEDIDOS, IDS E PREÇO ALEATÓRIAMENTO, DANDO SEQUENCIA A TABELA
for linha in range(5, 16):
    planilha1.cell(linha, 1).value = 'TESTE'
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

pedidos.save('nova_planilha.xlsx')


planilha = openpyxl.Workbook()
planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 1)

planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']

for linha in range(1, 11):
    planilha1.cell(linha, 1).value = 'TESTE'
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

for linha in range(1, 11):
    planilha2.cell(linha, 1).value = f'Lucas {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 2).value = f'Henrique {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 3).value = f'zézinho {linha} {round(uniform(10, 100), 2)}'

planilha.save('nova_planilha.xlsx')
